"""
Robô Playwright para a plataforma Track N' Me (conta "Broquel Rastreamento").

Reescrita a partir da lógica de `referencia_legado/broquel_bot.py`, com estas
mudanças (já fechadas na conversa):

    1. Login AUTOMÁTICO via usuário/senha da config (o legado fazia login
       manual — vira exceção, não regra).
    2. Handoff para o humano: se o login automático falhar, abre o
       navegador visível e pausa, esperando login manual — nunca falha a
       execução inteira silenciosamente.
    3. Detecção de sessão caída no meio da fila: cada ação (`abrir_incidente`,
       `concluir_incidente`, etc.) deve levantar
       `integrations.playwright_utils.SessaoCaidaError` ao detectar queda de
       sessão, para que `processar_fila` pause a fila inteira em vez de
       gastar tentativas por item.
    4. Duas ações de negócio, mapeadas nos grupos do motor de regras:
         abrir_incidente(placa, cliente)                       -> Grupo 1 (offline > 48h)
         concluir_incidente(placa, motivo, numero_incidente=None) -> Grupo 2 (voltou a comunicar)
    5. Download dos relatórios (Incidentes, Rastreadores Ativos) em Excel,
       antes de tudo isso (Fase A do pipeline).

Seletores confirmados por captura Playwright ao vivo (2026-08-04):
    Login: input[name="email"], input[name="password"], botão "LOG IN".
    Sinal de logado: mudança de URL (saiu de /login) — ver docstring de
        `SELETOR_PAGINA_RASTREAR_CARREGADA` abaixo pro motivo.
    Popup pós-login (release notes): div.modalRelease, botão "Close".
    Exportar Incidentes: nav "Operador" -> "Usar filtro avançado" ->
        Marca "PUMA NORDESTE" + Tipo "Sem comunicação por 48 horas" + Status
        ("Aberto" e "Em progresso", duas exportações mescladas em um
        arquivo) -> "Filtrar" -> "Exportar" -> popup "Fazer Download".
    Exportar Rastreadores Ativos: nav "Relatórios" -> Marca "PUMA NORDESTE"
        + Status "Ativo" (#mui-component-select-status) -> "Filtrar" ->
        "Exportar" -> popup "Fazer Download".
"""

import re
import sys
from dataclasses import dataclass
from pathlib import Path

import httpx
import openpyxl
from playwright.async_api import (
    Browser,
    BrowserContext,
    Page,
    Playwright,
    async_playwright,
    TimeoutError as PlaywrightTimeoutError,
)

from config import manager
from integrations.playwright_utils import SessaoCaidaError, processar_fila

URL_LOGIN = "https://broquelrastreamento.tracknme.com.br/monitoring/login"
URL_TRACKBACK = "https://broquelrastreamento.tracknme.com.br/monitoring/trackback"

# ---------------------------------------------------------------------------
# Caminho HTTP puro (achado 2026-08-19: sem captcha, dá pra logar e operar
# sem navegador nenhum -- ver `_handoff/investigacao_lag_relatorio_
# tracknme.md` pro método completo de descoberta). Cobre só abrir/concluir
# incidente por enquanto; download de relatórios continua via Playwright.
# ---------------------------------------------------------------------------

URL_BASE_API = "https://broquelrastreamento.tracknme.com.br/api"
NOME_MARCA_ALVO = "PUMA NORDESTE"
TIPO_SEM_COMUNICACAO_API = "NO_COMMUNICATION_48HS"
URL_INDICADORES = "https://broquelrastreamento.tracknme.com.br/monitoring/indicators"

SELETOR_CAMPO_EMAIL = 'input[name="email"]'
SELETOR_CAMPO_SENHA = 'input[name="password"]'
SELETOR_POPUP_RELEASE = "div.modalRelease"

# Sinal de que a página de "Rastrear" (trackback) carregou — NÃO existe na
# página de destino pós-login ("/monitoring/indicators", confirmado via
# screenshot em 2026-08-04: 0 elementos no DOM). O sinal de login
# bem-sucedido usado abaixo é a própria mudança de URL (saiu de "/login"),
# que é página-agnóstica; este seletor só é usado depois, ao navegar
# explicitamente para URL_TRACKBACK (abrir_incidente/concluir_incidente).
SELETOR_PAGINA_RASTREAR_CARREGADA = "#autoSuggest"

TIMEOUT_LOGIN_AUTOMATICO_MS = 15_000
TIMEOUT_LOGIN_MANUAL_MS = 300_000  # 5 minutos, mesmo timeout do legado


async def _fechar_popup_release(page: Page) -> None:
    """Fecha o popup de release notes pós-login, se aparecer (mesma lógica
    do legado — nem sempre aparece, por isso o timeout curto)."""
    try:
        await page.wait_for_selector(SELETOR_POPUP_RELEASE, state="visible", timeout=5_000)
    except PlaywrightTimeoutError:
        return

    for seletor in ['button[aria-label="Close"]', 'button:has-text("Close")', "div.modalRelease button"]:
        try:
            await page.click(seletor, timeout=2_000)
            return
        except PlaywrightTimeoutError:
            continue
    await page.keyboard.press("Escape")


async def _instalar_handler_popup_release(page: Page) -> None:
    """Registra um handler automático do Playwright (`add_locator_handler`)
    que fecha o popup de release notes sempre que ele bloquear alguma
    ação -- confirmado ao vivo que ele aparece em momentos imprevisíveis
    (timing variado, não é "só depois do login" nem "só depois de um
    clique específico"), então checar em pontos fixos do fluxo não é
    confiável o suficiente. Idempotente -- seguro chamar várias vezes na
    mesma página."""
    if getattr(page, "_popup_release_handler_instalado", False):
        return
    await page.add_locator_handler(page.locator(SELETOR_POPUP_RELEASE), lambda: _fechar_popup_release(page))
    page._popup_release_handler_instalado = True


async def _limpar_cache_navegador(page: Page) -> None:
    """Limpa o cache do Chromium via CDP antes de recarregar. O Track N'
    Me às vezes trava num estado ruim (login não confirma, telas não
    terminam de carregar) que um reload comum não resolve -- confirmado
    pelo usuário (2026-08-04) que o login manual funciona normalmente
    quando isso acontece, ou seja, é estado/cache do navegador
    automatizado que trava, não a conta nem o site em si."""
    client = await page.context.new_cdp_session(page)
    await client.send("Network.clearBrowserCache")


async def _fazer_login_automatico(page: Page) -> bool:
    """Preenche usuário/senha da config e confirma o sinal de login.
    Retorna False (sem levantar exceção) se não conseguir dentro do
    timeout — quem chama decide fazer o handoff pro humano.

    Se a primeira tentativa travar, limpa o cache do navegador e tenta
    mais uma vez antes de desistir (ver `_limpar_cache_navegador`)."""
    cfg = manager.carregar_config()["tracknme"]

    async def _tentar() -> bool:
        try:
            await page.goto(URL_LOGIN)
            await page.fill(SELETOR_CAMPO_EMAIL, cfg["usuario"])
            await page.fill(SELETOR_CAMPO_SENHA, cfg["senha"])
            await page.get_by_role("button", name="LOG IN").click()
            await page.wait_for_url(lambda url: "/login" not in url, timeout=TIMEOUT_LOGIN_AUTOMATICO_MS)
            return True
        except PlaywrightTimeoutError:
            return False

    if await _tentar():
        await _fechar_popup_release(page)
        return True

    await _limpar_cache_navegador(page)
    if await _tentar():
        await _fechar_popup_release(page)
        return True

    return False


async def abrir_navegador_autenticado(playwright: Playwright) -> tuple[Browser, BrowserContext]:
    """Ponto de entrada de autenticação: tenta login automático (headless,
    rápido); se falhar, faz o handoff pro humano — reabre visível e espera
    o login manual, sem levantar exceção nem derrubar a execução.
    """
    browser = await playwright.chromium.launch(headless=True)
    context = await browser.new_context()
    page = await context.new_page()

    if await _fazer_login_automatico(page):
        return browser, context

    await context.close()
    await browser.close()

    browser = await playwright.chromium.launch(headless=False)
    context = await browser.new_context()
    page = await context.new_page()
    await page.goto(URL_LOGIN)
    # TODO (etapa de integração com a UI): sinalizar aqui o estado
    # "faça o login manualmente e confirme" em vez de só esperar a URL
    # mudar — depende de orchestrator/pipeline.py e ui/app.py existirem.
    await page.wait_for_url(lambda url: "/login" not in url, timeout=TIMEOUT_LOGIN_MANUAL_MS)
    await _fechar_popup_release(page)
    return browser, context


async def testar_login() -> bool:
    """Usado pelo botão "Testar conexão" (`config.manager.testar_conexao`)
    — reaproveita `_fazer_login_automatico`, só a tentativa headless (sem
    o handoff pro humano de `abrir_navegador_autenticado`, que abriria um
    navegador visível só pra testar credencial). Sempre fecha
    browser/context, mesmo se o login falhar."""
    async with async_playwright() as playwright:
        browser = await playwright.chromium.launch(headless=True)
        context = await browser.new_context()
        page = await context.new_page()
        try:
            return await _fazer_login_automatico(page)
        finally:
            await context.close()
            await browser.close()


# ---------------------------------------------------------------------------
# baixar_relatorios()
# ---------------------------------------------------------------------------

_STATUS_INCIDENTES = ("Aberto", "Em progresso")


def _diretorio_downloads() -> Path:
    """Pasta `downloads/`, resolvida ao lado do código-fonte em dev; quando
    empacotado, em `%LOCALAPPDATA%\\ConsolidacaoTrackNMe\\downloads`
    (`config.manager._diretorio_dados_local`) — fixo por máquina,
    independente de qual pasta de versão do `.exe` está rodando (Fase 1,
    Launcher)."""
    if getattr(sys, "frozen", False):
        base = manager._diretorio_dados_local()
    else:
        base = Path(__file__).resolve().parent.parent
    return base / "downloads"


async def _selecionar_dropdown(page: Page, rotulo_regex: str, opcao: str) -> None:
    """Seleciona uma opção num dropdown MUI cujo rótulo muda de acordo com
    o valor atual (ex: "Status" -> "Status Aberto") — por isso usa regex de
    prefixo em vez de nome exato."""
    await page.get_by_role("button", name=re.compile(rotulo_regex)).click()
    await page.get_by_role("option", name=opcao, exact=True).click()


async def _exportar_e_baixar(page: Page):
    """Clicar "Exportar" só inicia o processamento assíncrono do relatório
    (confirmado ao vivo, 2026-08-04: o clique já dispara um evento de
    download próprio, um placeholder vazio — por isso NÃO pode estar
    dentro do mesmo `expect_download` do download real). Espera o modal
    "Download de Relatório" ficar pronto (pode levar alguns segundos) e só
    então clica em "Fazer Download", que abre um popup e dispara o
    download de verdade.
    """
    await page.get_by_role("button", name="Exportar").click()

    botao_download = page.get_by_role("button", name="Fazer Download")
    await botao_download.wait_for(state="visible", timeout=60_000)

    async with page.expect_download() as download_info:
        async with page.expect_popup() as popup_info:
            await botao_download.click()
        popup = await popup_info.value
    download = await download_info.value
    await popup.close()
    return download


def _verificar_sessao(page: Page) -> None:
    if "/login" in page.url:
        raise SessaoCaidaError("Sessão do Track N' Me caiu (redirecionado para /login)")


async def _garantir_autenticado(page: Page) -> None:
    """`processar_fila` cria páginas novas por worker dentro do mesmo
    `BrowserContext`, e a MESMA página é reaproveitada pra vários itens
    dentro de um worker. Confirmado por teste ao vivo (2026-08-04): uma
    página nova (em branco) é redirecionada pro login ao navegar direto
    pro dashboard, mesmo com os cookies do contexto válidos — o app guarda
    estado de sessão em algo por-página (ex: sessionStorage), não só
    cookie. Por isso: checa se ESSA página já está autenticada antes de
    refazer login (refazer sempre quebra a 2ª+ chamada na mesma página,
    que já teria saído do formulário de login).
    """
    await _instalar_handler_popup_release(page)
    await page.goto(URL_INDICADORES)
    if "/login" not in page.url:
        # já autenticada -- ainda assim, o popup de release notes pode
        # reaparecer a cada navegação pra /indicators (não é "uma vez só
        # por sessão"), então precisa fechar de novo antes de prosseguir.
        await _fechar_popup_release(page)
        return

    if not await _fazer_login_automatico(page):
        raise SessaoCaidaError("Não foi possível autenticar a página do worker")


async def _exportar_incidentes_por_status(page: Page, status: str) -> Path:
    await _garantir_autenticado(page)
    await page.get_by_text("headset_micOperador").click()
    await page.get_by_role("button", name="Usar filtro avançado").click()
    await page.get_by_role("textbox", name="Marca").click()
    await page.get_by_text("PUMA NORDESTE").click()
    await _selecionar_dropdown(page, r"^Tipo", "Sem comunicação por 48 horas")
    await _selecionar_dropdown(page, r"^Status", status)
    await page.get_by_role("button", name="Filtrar").click()
    download = await _exportar_e_baixar(page)
    _verificar_sessao(page)

    caminho = _diretorio_downloads() / f"_tmp_incidentes_{status.lower().replace(' ', '_')}.xlsx"
    await download.save_as(str(caminho))
    return caminho


async def _exportar_rastreadores_ativos(page: Page) -> Path:
    await _garantir_autenticado(page)
    await page.get_by_role("button", name="Relatórios").click()
    await page.get_by_role("textbox", name="Marca").click()
    await page.get_by_text("PUMA NORDESTE").click()
    await page.locator("#mui-component-select-status").click()
    await page.get_by_role("option", name="Ativo", exact=True).click()
    await page.get_by_role("button", name="Filtrar").click()
    download = await _exportar_e_baixar(page)
    _verificar_sessao(page)

    caminho = _diretorio_downloads() / "rastreadores_ativos.xlsx"
    await download.save_as(str(caminho))
    return caminho


async def _acao_download(page: Page, item: tuple[str, str | None]) -> Path:
    tipo, status = item
    if tipo == "incidentes":
        return await _exportar_incidentes_por_status(page, status)
    return await _exportar_rastreadores_ativos(page)


def ler_linhas_xlsx(caminho: Path) -> list[list]:
    """Lê todas as linhas não vazias de um .xlsx, sem confiar no atributo
    `<dimension>` do XML interno. O relatório do Track N' Me não preenche
    esse atributo corretamente: com `read_only=True` o openpyxl enxerga
    1 linha/1 coluna só, e mesmo sem `read_only` o `max_row` reportado vem
    inflado (ex: 90000) com a maior parte das linhas vazia. Por isso
    carregamos sem `read_only` e filtramos linhas totalmente vazias.

    Pública porque qualquer módulo que precisar ler `rastreadores_ativos.xlsx`
    (ou outro relatório baixado por este bot) deve reaproveitar esta função
    em vez de abrir o arquivo com `openpyxl` diretamente."""
    wb = openpyxl.load_workbook(caminho, data_only=True)
    try:
        return [
            list(linha)
            for linha in wb.active.iter_rows(values_only=True)
            if any(valor is not None for valor in linha)
        ]
    finally:
        wb.close()


def _mesclar_xlsx(caminhos: list[Path], caminho_destino: Path) -> None:
    """Mescla vários .xlsx com o mesmo cabeçalho em um único arquivo (usado
    pra juntar os exports de Incidentes "Aberto" + "Em progresso")."""
    cabecalho: list | None = None
    linhas: list[list] = []
    for caminho in caminhos:
        valores = ler_linhas_xlsx(caminho)
        if not valores:
            continue
        if cabecalho is None:
            cabecalho = list(valores[0])
        linhas.extend(valores[1:])

    wb_saida = openpyxl.Workbook()
    ws_saida = wb_saida.active
    if cabecalho is not None:
        ws_saida.append(cabecalho)
    for linha in linhas:
        ws_saida.append(linha)
    wb_saida.save(caminho_destino)


async def baixar_relatorios() -> dict:
    """Baixa Incidentes (status "Aberto" + "Em progresso", mesclados num só
    arquivo) e Rastreadores Ativos (status "Ativo"), sempre filtrando Marca
    "PUMA NORDESTE". Salva em `downloads/`, sempre sobrescrevendo (espelho,
    sem acumular histórico). Retorna os caminhos locais dos dois arquivos.

    Sempre roda as exportações em sequência (`num_workers=1` fixo,
    diferente do resto dos robôs Playwright, onde N workers é configurável).
    Confirmado ao vivo (2026-08-04): com exportações concorrentes sob a
    mesma conta, o backend do Track N' Me às vezes serve o arquivo de um
    relatório errado pro "Fazer Download" de outro (ex: rastreadores_ativos
    saindo com o cabeçalho de incidentes) — silencioso, sem erro, sem
    retry possível. Como são só 3 itens, o custo de rodar sequencial é
    baixo perto do risco de misturar dados errados.
    """
    _diretorio_downloads().mkdir(parents=True, exist_ok=True)

    itens: list[tuple[str, str | None]] = [
        ("incidentes", status) for status in _STATUS_INCIDENTES
    ] + [("rastreadores_ativos", None)]

    async with async_playwright() as playwright:
        browser, context = await abrir_navegador_autenticado(playwright)
        try:
            resultados = await processar_fila(context, itens, _acao_download, num_workers=1)
        finally:
            await context.close()
            await browser.close()

    falhas = [r for r in resultados if not r.sucesso]
    if falhas:
        raise RuntimeError(f"Falha ao baixar relatório(s): {[r.item for r in falhas]}")

    caminhos_incidentes = [r.resultado for r in resultados if r.item[0] == "incidentes"]
    caminho_incidentes = _diretorio_downloads() / "incidentes.xlsx"
    _mesclar_xlsx(caminhos_incidentes, caminho_incidentes)
    for caminho_bruto in caminhos_incidentes:
        caminho_bruto.unlink(missing_ok=True)

    caminho_rastreadores = next(
        r.resultado for r in resultados if r.item[0] == "rastreadores_ativos"
    )

    return {
        "incidentes": str(caminho_incidentes),
        "rastreadores_ativos": str(caminho_rastreadores),
    }


# ---------------------------------------------------------------------------
# abrir_incidente()
# ---------------------------------------------------------------------------

SELETOR_DROPDOWN_LISTBOX = 'div[role="tooltip"] ul[role="listbox"]'
SELETOR_ITENS_DROPDOWN = 'div[role="tooltip"] ul[role="listbox"] li[role="option"]'
TIPO_INCIDENTE_SEM_COMUNICACAO = "Sem comunicação por 48 horas"


class IncidenteDuplicadoError(RuntimeError):
    """`abrir_incidente` detectou que o Track N' Me rejeitou a criação
    porque a placa já tem um incidente aberto (confirmado ao vivo,
    2026-08-04: o clique em "Criar" não navega pra "Detalhes do
    incidente", só fecha o modal e volta pro painel do veículo). Ainda
    assim é um cenário que NÃO deveria acontecer em uso normal -- o motor
    de regras só chama `abrir_incidente` pra placas sem incidente aberto
    (ver `docs/prompt_claude_code.md`) -- então tratar como falha normal de
    item (retry 2 rounds + sinalizar pro atendente, decisão #8), não como
    `SessaoCaidaError`."""


async def _garantir_pagina_trackback(page: Page) -> None:
    """Mesma ideia de `_garantir_autenticado`, mas para `URL_TRACKBACK`
    (de onde `abrir_incidente`/`concluir_incidente` partem) em vez de
    `URL_INDICADORES`. O campo `#autoSuggest` começa OCULTO nesta página
    (confirmado ao vivo) -- só aparece depois de clicar no ícone de busca
    (`[data-testid="search"]`), igual o legado já tratava em
    `preparar_pagina`."""
    await _instalar_handler_popup_release(page)
    await page.goto(URL_TRACKBACK)
    if "/login" in page.url:
        if not await _fazer_login_automatico(page):
            raise SessaoCaidaError("Não foi possível autenticar a página do worker")
        await page.goto(URL_TRACKBACK)

    await _fechar_popup_release(page)

    if not await page.locator(SELETOR_PAGINA_RASTREAR_CARREGADA).is_visible():
        await page.click('[data-testid="search"]', timeout=5_000)
        await page.wait_for_selector(SELETOR_PAGINA_RASTREAR_CARREGADA, state="visible", timeout=15_000)


async def _buscar_e_selecionar_veiculo(page: Page, placa: str, cliente: str) -> bool:
    """Busca `placa` no autocomplete, filtra os resultados pelo `cliente`
    (pode haver placas parecidas de clientes diferentes) e clica no
    veículo certo na lista lateral. Retorna False se não achar (placa
    inválida/inexistente ou contrato inativo) -- não é uma falha técnica,
    por isso não levanta exceção aqui.

    O clique na lista lateral usa o texto combinado ícone+placa
    ("location_on<placa>...", confirmado ao vivo na Sessão 4 de captura,
    mesmo padrão MUI já usado em `_exportar_incidentes_por_status` pro nav
    "headset_micOperador") casando só o prefixo -- o sufixo de status
    (`Parado` no momento da captura) é dado de posição em tempo real do
    veículo e não deveria ser hardcoded.
    """
    async def _digitar_placa() -> None:
        await page.click("#autoSuggest", timeout=3_000)
        await page.fill("#autoSuggest", "")
        await page.type("#autoSuggest", placa, delay=80)
        await page.wait_for_timeout(800)
        await page.wait_for_selector(SELETOR_DROPDOWN_LISTBOX, state="visible", timeout=10_000)

    try:
        await _digitar_placa()
    except PlaywrightTimeoutError:
        return False

    itens = await page.query_selector_all(SELETOR_ITENS_DROPDOWN)
    cliente_normalizado = cliente.strip().lower()
    indices_validos = [
        idx
        for idx, item in enumerate(itens)
        if cliente_normalizado in ((await item.text_content()) or "").lower()
    ]
    if not indices_validos:
        return False

    for tentativa, idx in enumerate(indices_validos):
        if tentativa > 0:
            try:
                await _digitar_placa()
            except PlaywrightTimeoutError:
                continue
            itens = await page.query_selector_all(SELETOR_ITENS_DROPDOWN)
            if idx >= len(itens):
                continue

        await itens[idx].click()
        await page.wait_for_timeout(1_000)
        try:
            await page.get_by_text(re.compile(f"^location_on{re.escape(placa)}")).click(timeout=8_000)
            return True
        except PlaywrightTimeoutError:
            continue

    return False


async def _clicar_voltar(page: Page) -> None:
    """O botão "Voltar" (usado em `abrir_incidente` e `concluir_incidente`)
    tem o texto combinado ícone+label ("arrow_backVoltar", confirmado ao
    vivo) -- `get_by_text(..., exact=True)` não bate com isso, por isso
    usa regex parcial, mesmo padrão de `_buscar_e_selecionar_veiculo`."""
    await page.get_by_text(re.compile("Voltar")).click(timeout=5_000)


async def abrir_incidente(page: Page, placa: str, cliente: str) -> str:
    """Abre um incidente "Sem comunicação por 48 horas" pra `placa`
    (Grupo 1 do motor de regras: equipamento ativo, sem incidente aberto,
    offline > 48h). Fluxo confirmado ao vivo (Sessão 4 de captura,
    2026-08-04): busca placa no autocomplete -> filtra pela lista pelo
    `cliente` -> clica no veículo na lista lateral -> "Incidentes" ->
    dropdown de tipo (`#mui-component-select-description`, substitui o
    seletor frágil `get_by_label("", exact=True)`) -> "Sem comunicação por
    48 horas" -> "Criar".

    Falhas de negócio (placa não encontrada, modal não abriu, etc.)
    levantam `RuntimeError` -- não retornam uma string de erro -- porque
    `processar_fila` só reconhece falha de item via exceção (é isso que
    aciona o retry de 2 rounds da decisão de negócio já fechada). Queda de
    sessão continua levantando `SessaoCaidaError` à parte, que pausa a
    fila inteira em vez de gastar tentativas.
    """
    await _garantir_pagina_trackback(page)

    if not await _buscar_e_selecionar_veiculo(page, placa, cliente):
        raise RuntimeError(f"Veículo não encontrado ou contrato inativo (placa={placa})")

    try:
        await page.click('span[title="Criar Incidente"]', timeout=5_000)
    except PlaywrightTimeoutError:
        _verificar_sessao(page)
        raise RuntimeError("Botão 'Incidentes' não encontrado")

    try:
        await page.wait_for_selector("#btn-create-incident", state="visible", timeout=10_000)
    except PlaywrightTimeoutError:
        raise RuntimeError("Modal de criação de incidente não abriu")

    await page.click("#mui-component-select-description")
    await page.get_by_role("option", name=TIPO_INCIDENTE_SEM_COMUNICACAO, exact=True).click()
    await page.click("#btn-create-incident")

    try:
        await page.wait_for_selector("text=Detalhes do incidente", timeout=15_000)
    except PlaywrightTimeoutError:
        if await page.locator('span[title="Criar Incidente"]').is_visible():
            raise IncidenteDuplicadoError(
                f"Track N' Me rejeitou a criação -- provável incidente já aberto (placa={placa})"
            )
        raise RuntimeError("Incidente não confirmado (página de detalhes não abriu após 'Criar')")

    await _clicar_voltar(page)
    return "Incidente aberto"


async def _acao_abrir_incidente(page: Page, item: tuple[str, str]) -> str:
    placa, cliente = item
    return await abrir_incidente(page, placa, cliente)


# ---------------------------------------------------------------------------
# concluir_incidente()
# ---------------------------------------------------------------------------

SITUACAO_ANALISE_SISTEMA = "Análise pelo sistema"
TIPO_ACOMPANHAMENTO_INFORMACAO = "Informação"


class MultiplosIncidentesAbertosError(RuntimeError):
    """`concluir_incidente` achou mais de um incidente aberto pra mesma
    placa e não recebeu `numero_incidente` pra desambiguar -- confirmado
    ao vivo que a tela Operador tem um campo próprio de busca por número
    (`input[name="number"]`) pra esse caso (Sessão 5 de captura)."""


_MARCADORES_ERRO_NEGOCIO_ESPERADO = ("incidente já aberto", "Mais de um incidente aberto")


def eh_erro_de_negocio_esperado(mensagem: str | None) -> bool:
    """True se `mensagem` (tipicamente `ResultadoItem.erro`, já uma string
    -- `_executar_com_tentativas` só guarda `str(exceção)`, não o tipo)
    vem de `IncidenteDuplicadoError`/`MultiplosIncidentesAbertosError` --
    resultados de NEGÓCIO esperados (a API corretamente recusando
    duplicado/ambíguo), não falha técnica. Achado ao vivo 2026-08-19,
    teste de escala com 100 candidatos reais: 8% vieram de duplicado
    (não de instabilidade técnica) -- um circuit breaker que conta isso
    junto com falha técnica de verdade abortaria o caminho HTTP à toa.
    Usado por `orchestrator.pipeline._avaliar_circuit_breaker_tracknme_
    http` pra não contar esses casos no limiar."""
    if not mensagem:
        return False
    return any(marcador in mensagem for marcador in _MARCADORES_ERRO_NEGOCIO_ESPERADO)


async def _abrir_tela_operador(page: Page) -> None:
    """Navega até a tela "Operador" (fila de incidentes), de onde
    `concluir_incidente` busca o incidente aberto -- diferente de
    `abrir_incidente`, que parte de `URL_TRACKBACK`.

    Em todo este módulo evitamos `page.wait_for_load_state("networkidle")`
    de propósito: confirmado ao vivo que o app do Track N' Me nunca fica
    de fato "idle" (polling contínuo de Firebase Messaging visto no
    console), então essa espera trava até estourar o timeout genérico do
    Playwright (30s) sem erro claro. Preferimos pausas fixas curtas ou
    esperar por um elemento específico do próximo passo."""
    await _garantir_autenticado(page)
    await page.get_by_text("headset_micOperador").click()
    await page.wait_for_timeout(1_500)
    await _fechar_popup_release(page)


TIMEOUT_CARREGAMENTO_OPERADOR_MS = 45_000
TENTATIVAS_CARREGAMENTO_OPERADOR = 3


async def _preencher_filtro_incidente(page: Page, placa: str, numero_incidente: str | None) -> None:
    """Preenche os campos do filtro avançado -- extraído à parte porque
    precisa ser refeito do zero depois de um reload (ver
    `_buscar_incidente_aberto`)."""
    await page.get_by_role("button", name="Usar filtro avançado").click()
    await _fechar_popup_release(page)

    if numero_incidente:
        await page.fill('input[name="number"]', numero_incidente)
    else:
        await page.fill('input[name="licensePlate"]', placa)

    await _fechar_popup_release(page)
    await page.get_by_role("textbox", name="Marca").click()
    await page.get_by_text("PUMA NORDESTE").click()
    await _selecionar_dropdown(page, r"^Tipo", TIPO_INCIDENTE_SEM_COMUNICACAO)
    await _selecionar_dropdown(page, r"^Status", "Aberto")


async def _buscar_incidente_aberto(page: Page, placa: str, numero_incidente: str | None) -> str:
    """Usa "Usar filtro avançado" pra achar o incidente aberto de `placa`
    (ou de `numero_incidente`, quando há mais de um aberto na mesma placa)
    e clica no ícone "editar" da linha. Retorna o número do incidente
    encontrado.

    Confirmado pelo usuário (2026-08-04): a tela Operador às vezes trava
    carregando a listagem/formulário avançado -- problema conhecido de
    arquitetura do Track N' Me, não do nosso código, e que vai continuar
    acontecendo. Uma vez que a tela carrega de verdade (sem precisar de
    reload), ela costuma ficar estável -- por isso a estratégia é várias
    tentativas com timeout MODERADO + `page.reload()` entre elas, em vez
    de esperar um tempo enorme numa tentativa só (isso não resolveu ao
    vivo mesmo esperando 2 minutos)."""
    await _preencher_filtro_incidente(page, placa, numero_incidente)

    for tentativa in range(1, TENTATIVAS_CARREGAMENTO_OPERADOR + 1):
        try:
            await page.wait_for_selector(
                'button:has-text("Filtrar"):not([disabled])', timeout=TIMEOUT_CARREGAMENTO_OPERADOR_MS
            )
            break
        except PlaywrightTimeoutError:
            if tentativa == TENTATIVAS_CARREGAMENTO_OPERADOR:
                raise RuntimeError(
                    f"Tela Operador não terminou de carregar após {TENTATIVAS_CARREGAMENTO_OPERADOR} "
                    f"tentativas (placa={placa})"
                )
            await _limpar_cache_navegador(page)
            await page.reload()
            await page.wait_for_timeout(2_000)
            await _fechar_popup_release(page)
            await _preencher_filtro_incidente(page, placa, numero_incidente)

    await page.get_by_role("button", name="Filtrar").click()
    await page.wait_for_timeout(2_000)

    linhas = await page.query_selector_all("tbody tr")
    if not linhas:
        raise RuntimeError(f"Nenhum incidente aberto encontrado (placa={placa})")
    if len(linhas) > 1 and not numero_incidente:
        raise MultiplosIncidentesAbertosError(
            f"Mais de um incidente aberto pra placa={placa} -- informe numero_incidente"
        )

    linha = linhas[0]
    celulas = await linha.query_selector_all("td")
    numero = (await celulas[0].text_content() or "").strip() if celulas else ""

    botao_editar = await linha.query_selector("button")
    if botao_editar is None:
        raise RuntimeError(f"Ícone 'editar' não encontrado na linha do incidente (placa={placa})")
    await botao_editar.click(timeout=60_000)
    await page.wait_for_selector("text=Detalhes do incidente", timeout=45_000)
    await _fechar_popup_release(page)
    return numero


async def concluir_incidente(
    page: Page, placa: str, motivo: str, numero_incidente: str | None = None
) -> str:
    """Conclui o incidente "Sem comunicação por 48 horas" aberto de
    `placa` (Grupo 2 do motor de regras: veículo com incidente ativo que
    voltou a comunicar dentro de 48h). `motivo` é o texto do
    acompanhamento (vem de `rule_templates`, ex: REGRA_2 -- resolvido por
    quem chama, não por este módulo, pra não acoplar `tracknme_bot` ao
    Supabase). `numero_incidente` só é necessário quando a placa tem mais
    de um incidente aberto (ver `MultiplosIncidentesAbertosError`).

    Fluxo confirmado ao vivo (Sessão 5 de captura + reconhecimento
    adicional, 2026-08-04): tela Operador > "Usar filtro avançado" > busca
    por placa/número -> "editar" -> "Atribuir" -> "Alterar situação" para
    "Análise pelo sistema" -> aba "Acompanhamento" -> tipo "Informação" +
    `motivo` -> "Incluir" -> "Concluir Incidente" -> "Voltar".
    """
    await _abrir_tela_operador(page)
    numero = await _buscar_incidente_aberto(page, placa, numero_incidente)

    await page.get_by_role("button", name="Atribuir", exact=True).click()
    await page.wait_for_timeout(1_500)

    await page.click("#mui-component-select-status")
    await page.get_by_role("option", name=SITUACAO_ANALISE_SISTEMA, exact=True).click()
    await page.get_by_role("button", name="Alterar situação", exact=True).click()
    await page.wait_for_timeout(1_500)

    await page.get_by_text("Acompanhamento", exact=True).click()
    await page.click("#mui-component-select-type")
    await page.get_by_role("option", name=TIPO_ACOMPANHAMENTO_INFORMACAO, exact=True).click()
    await page.fill('textarea[name="comment"]', motivo)
    await page.get_by_role("button", name="Incluir", exact=True).click()
    await page.wait_for_timeout(1_500)

    await page.get_by_role("button", name="Concluir Incidente", exact=True).click()
    await page.wait_for_timeout(2_000)

    await _clicar_voltar(page)
    return f"Incidente {numero} concluído"


async def _acao_concluir_incidente(page: Page, item: tuple[str, str]) -> str:
    placa, motivo = item
    return await concluir_incidente(page, placa, motivo)


# ---------------------------------------------------------------------------
# Caminho HTTP puro -- abrir_incidente_http() / concluir_incidente_http()
# ---------------------------------------------------------------------------

@dataclass
class ContextoHttp:
    """Sessão HTTP autenticada, já resolvida pra marca certa (`NOME_MARCA_
    ALVO`) -- devolvida por `preparar_contexto_http`, reaproveitada por
    `abrir_incidente_http`/`concluir_incidente_http` (1 login por execução
    do pipeline, não por item). Quem chama é responsável por fechar
    `cliente` (`await contexto.cliente.aclose()`) quando terminar."""

    cliente: httpx.AsyncClient
    brand_id: int
    user_id: int


async def preparar_contexto_http() -> ContextoHttp:
    """Login 100% automático via HTTP puro -- Track N' Me não tem captcha
    (diferente do SGA), então não precisa de navegador nem de
    `storage_state`: 1 `POST` de login já autentica a sessão inteira.

    Corpo do login (`login`/`password`/`brand`/`persistent`/`type`) e o
    header `Authorization` (setado como default do cliente HTTP depois do
    login, não por request -- confirmado ao vivo que é assim que o app
    real faz, `axios.defaults.headers.common.Authorization = accessToken`,
    `accessToken` já vem com o prefixo `"Bearer "`) foram capturados do
    tráfego de rede real de um login automático de produção (2026-08-19)
    -- nunca documentados publicamente, só descobertos testando.

    Levanta `RuntimeError` se a conta tiver mais de 1 marca raiz (login
    ambíguo, nunca visto em produção até agora) ou se a árvore de marcas
    não tiver `NOME_MARCA_ALVO`.
    """
    cfg = manager.carregar_config()["tracknme"]
    cliente = httpx.AsyncClient(
        base_url=URL_BASE_API, headers={"Content-Type": "application/json"}, timeout=30.0
    )
    try:
        resposta_marcas = await cliente.get("/sessions/brands-by-login", params={"login": cfg["usuario"]})
        marcas_raiz = resposta_marcas.json().get("content", [])
        if len(marcas_raiz) != 1:
            raise RuntimeError(f"esperava exatamente 1 marca raiz pro login HTTP, veio {len(marcas_raiz)}")
        brand_raiz_id = marcas_raiz[0]["id"]

        resposta_login = await cliente.post(
            "/sessions",
            params={"essential": ""},
            json={
                "login": cfg["usuario"],
                "password": cfg["senha"],
                "brand": brand_raiz_id,
                "persistent": False,
                "type": "DEFAULT",
            },
        )
        if resposta_login.status_code >= 400:
            raise RuntimeError(f"login via HTTP falhou: status {resposta_login.status_code}")
        corpo_login = resposta_login.json()
        cliente.headers["Authorization"] = corpo_login["accessToken"]
        user_id = corpo_login["user"]["id"]

        resposta_arvore = await cliente.get(f"/v2/brands/tree/{brand_raiz_id}")
        marca_alvo = next(
            (
                m
                for m in resposta_arvore.json().get("content", [])
                if m.get("brandChildName") == NOME_MARCA_ALVO
            ),
            None,
        )
        if marca_alvo is None:
            raise RuntimeError(f"marca '{NOME_MARCA_ALVO}' não encontrada na árvore de marcas da conta")

        return ContextoHttp(cliente=cliente, brand_id=marca_alvo["brandChildId"], user_id=user_id)
    except Exception:
        await cliente.aclose()
        raise


async def _buscar_device_por_placa(contexto: ContextoHttp, placa: str, cliente_nome: str) -> dict | None:
    """Busca o device (rastreador) + vehicle de `placa` via
    `GET /v2/devices/detail`, filtrando pela marca certa (`contexto.
    brand_id`) e pelo `cliente_nome` (substring, mesma normalização de
    `_buscar_e_selecionar_veiculo` -- placas repetidas entre marcas/
    clientes diferentes são esperadas, é uma cooperativa multimarca).
    `None` = não encontrado (placa inválida/inexistente/contrato
    inativo), não é erro técnico -- mesma semântica de
    `_buscar_e_selecionar_veiculo` retornando `False`.

    Devolve o dict cru do endpoint (`id`=deviceId, `vehicleId`, `brandId`,
    `chassi`, `customerName`, ...). **Nunca usar `/v2/vehicles?
    licensePlate=` pra isso** -- esse endpoint só devolve 1 id, e usar o
    mesmo valor pra `deviceId`/`vehicleId` faz o incidente ficar invisível
    pro relatório baixado depois (achado real, confirmado com 3 testes
    reais -- ver `_handoff/investigacao_lag_relatorio_tracknme.md`).
    """
    resposta = await contexto.cliente.get(
        "/v2/devices/detail",
        params={"filter": "devices", "limit": 10, "page": 0, "status": "ACTIVE", "term": placa},
    )
    if resposta.status_code >= 400:
        raise RuntimeError(f"busca de device falhou: status {resposta.status_code} (placa={placa})")

    cliente_normalizado = cliente_nome.strip().lower()
    for candidato in resposta.json().get("content", []):
        if candidato.get("brandId") != contexto.brand_id:
            continue
        if cliente_normalizado in (candidato.get("customerName") or "").lower():
            return candidato
    return None


async def abrir_incidente_http(contexto: ContextoHttp, placa: str, cliente: str) -> str:
    """Equivalente HTTP de `abrir_incidente` -- mesmo contrato de retorno/
    exceções (retorna `"Incidente aberto"`, levanta `RuntimeError` se o
    veículo não for encontrado, `IncidenteDuplicadoError` se a API
    rejeitar por já existir incidente aberto pro device -- confirmado ao
    vivo, 2026-08-19: `POST .../operation/create` devolve `400` com
    `{"message": "já existe um incidente do tipo NO_COMMUNICATION_48HS
    aberto para esse dispositivo"}`, mensagem estável o bastante pra
    detectar por substring)."""
    device = await _buscar_device_por_placa(contexto, placa, cliente)
    if device is None:
        raise RuntimeError(f"Veículo não encontrado ou contrato inativo (placa={placa})")

    payload = {
        "brandId": contexto.brand_id,
        "deviceId": int(device["id"]),
        "vehicleId": int(device["vehicleId"]),
        "userOperatorId": contexto.user_id,
        "type": TIPO_SEM_COMUNICACAO_API,
        "observation": "",
        "returned": None,
    }
    resposta = await contexto.cliente.post("/v2/incidents/operation/create", json=payload)
    if resposta.status_code >= 400:
        mensagem = ""
        try:
            mensagem = str(resposta.json().get("message", ""))
        except Exception:  # noqa: BLE001 - corpo pode não ser JSON válido
            pass
        if "já existe" in mensagem and "aberto" in mensagem:
            raise IncidenteDuplicadoError(
                f"Track N' Me rejeitou a criação via HTTP -- incidente já aberto (placa={placa}): {mensagem}"
            )
        raise RuntimeError(f"criar incidente via HTTP falhou: status {resposta.status_code} (placa={placa})")

    corpo = resposta.json()
    if not corpo.get("id"):
        raise RuntimeError(f"criação via HTTP não devolveu id (placa={placa}): {corpo}")
    return "Incidente aberto"


async def _acao_abrir_incidente_http(contexto: ContextoHttp, item: tuple[str, str]) -> str:
    placa, cliente = item
    return await abrir_incidente_http(contexto, placa, cliente)


async def _buscar_incidente_aberto_http(contexto: ContextoHttp, placa: str) -> str:
    """Só usado quando `concluir_incidente_http` não recebe `numero_
    incidente` (fallback raro -- na prática o motor de regras sempre
    preenche `id`, igual o caminho Playwright já assume hoje). Usa
    `GET /v2/incidents/details`, que tem uma defasagem real de propagação
    conhecida (`_handoff/investigacao_lag_relatorio_tracknme.md`) -- não
    confiável pra "acabou de abrir", mas incidentes mais antigos (o caso
    normal aqui) já devem ter propagado."""
    resposta = await contexto.cliente.get(
        "/v2/incidents/details",
        params={
            "brandId": contexto.brand_id,
            "licensePlate": placa,
            "status": "OPEN",
            "type": TIPO_SEM_COMUNICACAO_API,
        },
    )
    if resposta.status_code >= 400:
        raise RuntimeError(f"busca de incidente aberto falhou: status {resposta.status_code} (placa={placa})")

    linhas = resposta.json().get("content", [])
    if not linhas:
        raise RuntimeError(f"Nenhum incidente aberto encontrado (placa={placa})")
    if len(linhas) > 1:
        raise MultiplosIncidentesAbertosError(
            f"Mais de um incidente aberto pra placa={placa} -- informe numero_incidente"
        )
    return str(linhas[0]["id"])


async def concluir_incidente_http(
    contexto: ContextoHttp, placa: str, motivo: str, numero_incidente: str | None = None
) -> str:
    """Equivalente HTTP de `concluir_incidente`. A conclusão real da tela
    NÃO é 1 chamada só -- é uma sequência de 4 (capturada ao vivo do
    tráfego de rede real de uma conclusão de produção, 2026-08-19):
    "Atribuir" -> "Alterar situação" -> "Acompanhamento" (`motivo`) ->
    "Concluir Incidente". Repetir só o último passo conclui o incidente,
    mas deixa `situation` vazia e sem nenhum comentário registrado --
    testado e confirmado incompleto perto do que a tela produz. Aborta
    com o erro do passo que falhar, sem tentar os seguintes (evita deixar
    o incidente pela metade sem sinalizar)."""
    numero = numero_incidente or await _buscar_incidente_aberto_http(contexto, placa)

    resposta_atribuir = await contexto.cliente.post(
        f"/v2/incidents/operation/assing/{numero}", json={"userOperatorId": contexto.user_id}
    )
    if resposta_atribuir.status_code >= 400:
        raise RuntimeError(f"atribuir operador falhou: status {resposta_atribuir.status_code} (incidente={numero})")

    resposta_situacao = await contexto.cliente.post(
        f"/v2/incidents/operation/situation/{numero}", json={"situation": SITUACAO_ANALISE_SISTEMA}
    )
    if resposta_situacao.status_code >= 400:
        raise RuntimeError(f"alterar situação falhou: status {resposta_situacao.status_code} (incidente={numero})")

    resposta_comentario = await contexto.cliente.post(
        f"/v2/incidents/operation/comment/{numero}",
        json={"comment": motivo, "type": TIPO_ACOMPANHAMENTO_INFORMACAO, "createdUserId": contexto.user_id},
    )
    if resposta_comentario.status_code >= 400:
        raise RuntimeError(f"registrar acompanhamento falhou: status {resposta_comentario.status_code} (incidente={numero})")

    resposta_resolver = await contexto.cliente.post(
        f"/v2/incidents/operation/resolved/{numero}", json={"data": {"loggedUser": str(contexto.user_id)}}
    )
    if resposta_resolver.status_code >= 400:
        raise RuntimeError(f"concluir incidente falhou: status {resposta_resolver.status_code} (incidente={numero})")

    return f"Incidente {numero} concluído"


async def _acao_concluir_incidente_http(contexto: ContextoHttp, item: tuple[str, str, str | None]) -> str:
    placa, motivo, numero_incidente = item
    return await concluir_incidente_http(contexto, placa, motivo, numero_incidente=numero_incidente)
